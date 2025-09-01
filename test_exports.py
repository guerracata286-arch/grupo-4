#!/usr/bin/env python3
"""
Test script to verify PDF and Excel export functionality
Run this after installing the required dependencies
"""

def test_imports():
    """Test that all required libraries can be imported"""
    try:
        import reportlab
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        print("✅ ReportLab imports successful")
    except ImportError as e:
        print(f"❌ ReportLab import failed: {e}")
        return False
    
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        print("✅ OpenPyXL imports successful")
    except ImportError as e:
        print(f"❌ OpenPyXL import failed: {e}")
        return False
    
    return True

def test_basic_pdf():
    """Test basic PDF generation"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        import io
        
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        p.drawString(100, 750, "Test PDF Generation")
        p.save()
        
        pdf_data = buffer.getvalue()
        buffer.close()
        
        if len(pdf_data) > 0:
            print("✅ Basic PDF generation successful")
            return True
        else:
            print("❌ PDF generation failed - no data")
            return False
    except Exception as e:
        print(f"❌ PDF generation failed: {e}")
        return False

def test_basic_excel():
    """Test basic Excel generation"""
    try:
        from openpyxl import Workbook
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        ws['A1'] = "Test Excel Generation"
        
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_data = buffer.getvalue()
        buffer.close()
        
        if len(excel_data) > 0:
            print("✅ Basic Excel generation successful")
            return True
        else:
            print("❌ Excel generation failed - no data")
            return False
    except Exception as e:
        print(f"❌ Excel generation failed: {e}")
        return False

if __name__ == "__main__":
    print("Testing export functionality...\n")
    
    success = True
    success &= test_imports()
    success &= test_basic_pdf()
    success &= test_basic_excel()
    
    if success:
        print("\n🎉 All tests passed! Export functionality should work correctly.")
        print("\nNext steps:")
        print("1. Start your Django development server")
        print("2. Navigate to the reports page (/reportes/)")
        print("3. Set your desired date range and filters")
        print("4. Click the 'Exportar PDF' or 'Exportar Excel' buttons")
        print("5. The files should download automatically")
    else:
        print("\n❌ Some tests failed. Please install the required dependencies:")
        print("pip install reportlab==4.0.7 openpyxl==3.1.2")
