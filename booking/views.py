from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth import logout, login
from django.contrib.auth.models import User
from .forms import ReservationForm, BlackoutForm, MaterialForm, InventoryForm, InventoryUpdateForm, CustomUserCreationForm, AdminUserCreationForm
from django.http import HttpResponse
from django.utils import timezone
from datetime import time, datetime, date
from django.db.models import Count, Sum, Q
from .models import Room, Material, RoomInventory, Reservation, ReservationItem, Blackout
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io

def is_library_admin(user):
    return user.is_authenticated and (user.is_staff or user.groups.filter(name='AdminBiblioteca').exists())

def index(request):
    # Redirect unauthenticated users to login
    if not request.user.is_authenticated:
        return redirect('login')
    
    # Redirect based on user type
    if request.user.is_staff or request.user.groups.filter(name='AdminBiblioteca').exists():
        # Admin users go to reports
        return redirect('reports')
    else:
        # Teachers and other users go to reservations
        return redirect('reservation_list')

@user_passes_test(lambda u: u.is_authenticated)
def reservation_create(request):
    materials = Material.objects.order_by('name')
    if request.method == "POST":
        form = ReservationForm(request.POST)
        items = []
        for m in materials:
            q = int(request.POST.get(f"qty_{m.id}", 0) or 0)
            if q > 0:
                items.append((m.id, q))
        if form.is_valid():
            room = form.cleaned_data["room"]
            date = form.cleaned_data["date"]
            start = form.cleaned_data["start_time"]
            end = form.cleaned_data["end_time"]

            # Validaciones simples (choque de reservas)
            exists = Reservation.objects.filter(room=room, date=date, start_time__lt=end, end_time__gt=start).exists()
            if exists:
                messages.error(request, "El salón ya está ocupado en ese horario.")
                return redirect('reservation_create')

            # Validación horario laboral
            if not (time(8,0) <= start < time(18,0) and time(8,0) < end <= time(18,0)):
                messages.error(request, "Horario permitido: 08:00 a 18:00.")
                return redirect('reservation_create')

            # Blackouts
            from datetime import datetime as _dt
            def _join(d,t): return _dt.combine(d,t)
            start_dt = _join(date, start); end_dt = _join(date, end)
            blackout_exists = Blackout.objects.filter(
                room__isnull=True, start_datetime__lt=end_dt, end_datetime__gt=start_dt
            ).exists() or Blackout.objects.filter(room=room, start_datetime__lt=end_dt, end_datetime__gt=start_dt).exists()
            if blackout_exists:
                messages.error(request, "Existe un bloqueo de agenda en ese horario (feriado/reunión).")
                return redirect('reservation_create')

            # Stock
            for mid, qty in items:
                inv = RoomInventory.objects.filter(room=room, material_id=mid).first()
                if not inv or inv.quantity < qty:
                    messages.error(request, "No hay stock suficiente de materiales para ese salón.")
                    return redirect('reservation_create')

            r = Reservation.objects.create(room=room, date=date, start_time=start, end_time=end, user=request.user)
            for mid, qty in items:
                ReservationItem.objects.create(reservation=r, material_id=mid, quantity=qty)
                inv = RoomInventory.objects.get(room=room, material_id=mid)
                inv.quantity -= qty; inv.save()

            # Create blackout for the reservation
            start_dt = _join(date, start)
            end_dt = _join(date, end)
            username = request.user.username
            Blackout.objects.create(
                room=room,
                start_datetime=start_dt,
                end_datetime=end_dt,
                reason=f"Reserva de {username}",
                created_by=request.user
            )

            messages.success(request, "Reserva creada con éxito.")
            return redirect('index')
    else:
        form = ReservationForm()
    return render(request, 'reservation_form.html', {'form': form, 'materials': materials})


def reservation_list(request):
    """List reservations - teachers see only their own, admins see all"""
    if request.user.is_authenticated:
        # Check if user is admin (staff or AdminBiblioteca group)
        if request.user.is_staff or request.user.groups.filter(name='AdminBiblioteca').exists():
            # Admins can see all reservations
            reservations = Reservation.objects.select_related('room', 'user').prefetch_related('items__material').order_by('-date', '-start_time')
        else:
            # Teachers (Docente group) and other users see only their own reservations
            reservations = Reservation.objects.filter(user=request.user).select_related('room', 'user').prefetch_related('items__material').order_by('-date', '-start_time')
    else:
        # Anonymous users see no reservations
        reservations = Reservation.objects.none()
    
    return render(request, 'reservations/list.html', {'reservations': reservations})

@user_passes_test(is_library_admin)
def blackout_list(request):
    # Only show administrative blackouts, not reservation-generated ones
    items = Blackout.objects.select_related('room').exclude(
        reason__startswith='Reserva de'
    ).order_by('-start_datetime')
    return render(request, 'blackouts/list.html', {'items': items})

@user_passes_test(is_library_admin)
def blackout_create(request):
    if request.method == "POST":
        form = BlackoutForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            
            # Cancel overlapping reservations before saving the blackout
            from datetime import datetime as _dt
            start_dt = obj.start_datetime
            end_dt = obj.end_datetime
            
            # Find overlapping reservations
            if obj.room:
                # Room-specific blackout
                overlapping_reservations = Reservation.objects.filter(
                    room=obj.room,
                    date=start_dt.date(),
                    start_time__lt=end_dt.time(),
                    end_time__gt=start_dt.time()
                )
            else:
                # Global blackout affects all rooms
                overlapping_reservations = Reservation.objects.filter(
                    date=start_dt.date(),
                    start_time__lt=end_dt.time(),
                    end_time__gt=start_dt.time()
                )
            
            cancelled_count = 0
            for reservation in overlapping_reservations:
                # Restore inventory for cancelled reservation
                for item in reservation.items.all():
                    inventory = RoomInventory.objects.get(
                        room=reservation.room, 
                        material=item.material
                    )
                    inventory.quantity += item.quantity
                    inventory.save()
                
                # Delete the reservation-generated blackout
                Blackout.objects.filter(
                    room=reservation.room,
                    reason=f"Reserva de {reservation.user.username}",
                    start_datetime=_dt.combine(reservation.date, reservation.start_time),
                    end_datetime=_dt.combine(reservation.date, reservation.end_time)
                ).delete()
                
                reservation.delete()
                cancelled_count += 1
            
            obj.save()
            
            if cancelled_count > 0:
                messages.success(request, f"Bloqueo creado. Se cancelaron {cancelled_count} reserva(s) que se solapaban.")
            else:
                messages.success(request, "Bloqueo creado.")
            return redirect('blackout_list')
    else:
        form = BlackoutForm()
    return render(request, 'blackouts/form.html', {'form': form, 'title': 'Nuevo bloqueo'})

@user_passes_test(is_library_admin)
def blackout_update(request, pk):
    obj = get_object_or_404(Blackout, pk=pk)
    if request.method == "POST":
        form = BlackoutForm(request.POST, instance=obj)
        if form.is_valid():
            updated_obj = form.save(commit=False)
            
            # Cancel overlapping reservations with the updated blackout
            from datetime import datetime as _dt
            start_dt = updated_obj.start_datetime
            end_dt = updated_obj.end_datetime
            
            # Find overlapping reservations
            if updated_obj.room:
                # Room-specific blackout
                overlapping_reservations = Reservation.objects.filter(
                    room=updated_obj.room,
                    date=start_dt.date(),
                    start_time__lt=end_dt.time(),
                    end_time__gt=start_dt.time()
                )
            else:
                # Global blackout affects all rooms
                overlapping_reservations = Reservation.objects.filter(
                    date=start_dt.date(),
                    start_time__lt=end_dt.time(),
                    end_time__gt=start_dt.time()
                )
            
            cancelled_count = 0
            for reservation in overlapping_reservations:
                # Restore inventory for cancelled reservation
                for item in reservation.items.all():
                    inventory = RoomInventory.objects.get(
                        room=reservation.room, 
                        material=item.material
                    )
                    inventory.quantity += item.quantity
                    inventory.save()
                
                # Delete the reservation-generated blackout
                Blackout.objects.filter(
                    room=reservation.room,
                    reason=f"Reserva de {reservation.user.username}",
                    start_datetime=_dt.combine(reservation.date, reservation.start_time),
                    end_datetime=_dt.combine(reservation.date, reservation.end_time)
                ).delete()
                
                reservation.delete()
                cancelled_count += 1
            
            updated_obj.save()
            
            if cancelled_count > 0:
                messages.success(request, f"Bloqueo actualizado. Se cancelaron {cancelled_count} reserva(s) que se solapaban.")
            else:
                messages.success(request, "Bloqueo actualizado.")
            return redirect('blackout_list')
    else:
        form = BlackoutForm(instance=obj)
    return render(request, 'blackouts/form.html', {'form': form, 'title': 'Editar bloqueo'})

@user_passes_test(is_library_admin)
def blackout_delete(request, pk):
    obj = get_object_or_404(Blackout, pk=pk)
    if request.method == "POST":
        obj.delete()
        messages.success(request, "Bloqueo eliminado.")
        return redirect('blackout_list')
    return render(request, 'blackouts/delete.html', {'obj': obj})

# Material Management Views
@user_passes_test(is_library_admin)
def material_list(request):
    materials = Material.objects.order_by('name')
    return render(request, 'materials/list.html', {'materials': materials})

@user_passes_test(is_library_admin)
def material_create(request):
    if request.method == "POST":
        form = MaterialForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Material creado exitosamente.")
            return redirect('material_list')
    else:
        form = MaterialForm()
    return render(request, 'materials/form.html', {'form': form, 'title': 'Nuevo Material'})

@user_passes_test(is_library_admin)
def material_update(request, pk):
    material = get_object_or_404(Material, pk=pk)
    if request.method == "POST":
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            messages.success(request, "Material actualizado exitosamente.")
            return redirect('material_list')
    else:
        form = MaterialForm(instance=material)
    return render(request, 'materials/form.html', {'form': form, 'title': 'Editar Material'})

@user_passes_test(is_library_admin)
def material_delete(request, pk):
    material = get_object_or_404(Material, pk=pk)
    if request.method == "POST":
        try:
            material.delete()
            messages.success(request, "Material eliminado exitosamente.")
        except Exception as e:
            messages.error(request, f"No se puede eliminar el material: {str(e)}")
        return redirect('material_list')
    return render(request, 'materials/delete.html', {'material': material})

# Inventory Management Views
@user_passes_test(is_library_admin)
def inventory_list(request):
    inventory = RoomInventory.objects.select_related('room', 'material').order_by('room__code', 'material__name')
    rooms = Room.objects.order_by('code')
    materials = Material.objects.order_by('name')
    return render(request, 'inventory/list.html', {
        'inventory': inventory,
        'rooms': rooms,
        'materials': materials
    })

@user_passes_test(is_library_admin)
def inventory_create(request):
    if request.method == "POST":
        form = InventoryForm(request.POST)
        if form.is_valid():
            room = form.cleaned_data['room']
            material = form.cleaned_data['material']
            # Check if inventory already exists
            existing = RoomInventory.objects.filter(room=room, material=material).first()
            if existing:
                messages.error(request, f"Ya existe inventario para {material.name} en salón {room.code}")
                return render(request, 'inventory/form.html', {'form': form, 'title': 'Agregar Inventario'})
            form.save()
            messages.success(request, "Inventario agregado exitosamente.")
            return redirect('inventory_list')
    else:
        form = InventoryForm()
    return render(request, 'inventory/form.html', {'form': form, 'title': 'Agregar Inventario'})

@user_passes_test(is_library_admin)
def inventory_update(request, pk):
    inventory = get_object_or_404(RoomInventory, pk=pk)
    if request.method == "POST":
        form = InventoryUpdateForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            quantity = form.cleaned_data['quantity']
            
            if action == 'add':
                inventory.quantity += quantity
            elif action == 'remove':
                new_qty = inventory.quantity - quantity
                if new_qty < 0:
                    messages.error(request, "No se puede quitar más cantidad de la disponible.")
                    return render(request, 'inventory/update.html', {'form': form, 'inventory': inventory})
                inventory.quantity = new_qty
            elif action == 'set':
                inventory.quantity = quantity
            
            inventory.save()
            messages.success(request, f"Inventario actualizado: {inventory.material.name} en salón {inventory.room.code}")
            return redirect('inventory_list')
    else:
        form = InventoryUpdateForm()
    return render(request, 'inventory/update.html', {'form': form, 'inventory': inventory})

@user_passes_test(is_library_admin)
def inventory_delete(request, pk):
    inventory = get_object_or_404(RoomInventory, pk=pk)
    if request.method == "POST":
        inventory.delete()
        messages.success(request, "Inventario eliminado exitosamente.")
        return redirect('inventory_list')
    return render(request, 'inventory/delete.html', {'item': inventory})

def custom_logout(request):
    """Custom logout view that properly clears session and forces redirect"""
    logout(request)
    response = redirect('/')
    # Clear all cookies related to authentication
    response.delete_cookie('sessionid')
    response.delete_cookie('csrftoken')
    # Add cache control headers to prevent caching
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def user_register(request):
    """User registration view"""
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, f'Cuenta creada exitosamente para {user.username}!')
            return redirect('index')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/register.html', {'form': form})


@user_passes_test(is_library_admin)
def user_list(request):
    """List all users - only accessible to admins"""
    users = User.objects.select_related().prefetch_related('groups').order_by('username')
    return render(request, 'users/list.html', {'users': users})


@user_passes_test(is_library_admin)
def user_create(request):
    """Create new user - only accessible to admins"""
    if request.method == 'POST':
        form = AdminUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'Usuario {user.username} creado exitosamente.')
            return redirect('user_list')
    else:
        form = AdminUserCreationForm()
    return render(request, 'users/form.html', {'form': form, 'title': 'Nuevo Usuario'})


@user_passes_test(is_library_admin)
def reports_view(request):
    """Reports view with date range and room filters"""
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    room_filter = request.GET.get('room')
    
    # Default to current month if no dates provided
    if not start_date or not end_date:
        today = date.today()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    
    # Parse dates
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "Formato de fecha inválido")
        today = date.today()
        start_date_obj = today.replace(day=1)
        end_date_obj = today
        start_date = start_date_obj.strftime('%Y-%m-%d')
        end_date = end_date_obj.strftime('%Y-%m-%d')
    
    # Base queryset for reservations in date range
    reservations_qs = Reservation.objects.filter(
        date__gte=start_date_obj,
        date__lte=end_date_obj
    )
    
    # Apply room filter if specified
    if room_filter:
        reservations_qs = reservations_qs.filter(room_id=room_filter)
    
    # Report 1: Reservations by room (count)
    room_stats = reservations_qs.values(
        'room__code'
    ).annotate(
        reservation_count=Count('id')
    ).order_by('room__code')
    
    # Report 2: Materials requested (sum by type)
    material_stats = ReservationItem.objects.filter(
        reservation__in=reservations_qs
    ).values(
        'material__name'
    ).annotate(
        total_quantity=Sum('quantity')
    ).order_by('material__name')
    
    # Get all rooms for filter dropdown
    rooms = Room.objects.order_by('code')
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'room_filter': room_filter,
        'room_stats': room_stats,
        'material_stats': material_stats,
        'rooms': rooms,
        'total_reservations': reservations_qs.count(),
        'date_range_display': f"{start_date_obj.strftime('%d/%m/%Y')} - {end_date_obj.strftime('%d/%m/%Y')}"
    }
    
    return render(request, 'reports/dashboard.html', context)


@user_passes_test(is_library_admin)
def export_reports_pdf(request):
    """Export reports data to PDF"""
    # Get the same filter parameters as reports_view
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    room_filter = request.GET.get('room')
    
    # Default to current month if no dates provided
    if not start_date or not end_date:
        today = date.today()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    
    # Parse dates
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        today = date.today()
        start_date_obj = today.replace(day=1)
        end_date_obj = today
    
    # Get the same data as reports_view
    reservations_qs = Reservation.objects.filter(
        date__gte=start_date_obj,
        date__lte=end_date_obj
    )
    
    if room_filter:
        reservations_qs = reservations_qs.filter(room_id=room_filter)
    
    room_stats = reservations_qs.values(
        'room__code'
    ).annotate(
        reservation_count=Count('id')
    ).order_by('room__code')
    
    material_stats = ReservationItem.objects.filter(
        reservation__in=reservations_qs
    ).values(
        'material__name'
    ).annotate(
        total_quantity=Sum('quantity')
    ).order_by('material__name')
    
    # Check if there's data to export
    if not reservations_qs.exists():
        messages.error(request, "No hay datos para exportar en el período seleccionado.")
        return redirect('reports')
    
    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_biblioteca_{start_date}_{end_date}.pdf"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Title
    title = Paragraph("Reporte de Biblioteca", title_style)
    elements.append(title)
    
    # Date range
    date_range = Paragraph(
        f"Período: {start_date_obj.strftime('%d/%m/%Y')} - {end_date_obj.strftime('%d/%m/%Y')}",
        styles['Normal']
    )
    elements.append(date_range)
    elements.append(Spacer(1, 20))
    
    # Summary stats
    summary_data = [
        ['Métrica', 'Valor'],
        ['Total de reservas', str(reservations_qs.count())],
        ['Salones utilizados', str(len(room_stats))],
        ['Tipos de materiales', str(len(material_stats))]
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Room statistics
    elements.append(Paragraph("Reservas por Salón", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    if room_stats:
        room_data = [['Código de Salón', 'Cantidad de Reservas']]
        for stat in room_stats:
            room_data.append([
                f"Salón {stat['room__code']}",
                str(stat['reservation_count'])
            ])
        
        room_table = Table(room_data)
        room_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(room_table)
    else:
        elements.append(Paragraph("No hay datos de reservas para el período seleccionado.", styles['Normal']))
    
    elements.append(Spacer(1, 30))
    
    # Material statistics
    elements.append(Paragraph("Materiales Solicitados", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    if material_stats:
        material_data = [['Material', 'Cantidad Total Solicitada']]
        for stat in material_stats:
            material_data.append([
                stat['material__name'],
                str(stat['total_quantity'])
            ])
        
        material_table = Table(material_data)
        material_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(material_table)
    else:
        elements.append(Paragraph("No hay datos de materiales para el período seleccionado.", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


@user_passes_test(is_library_admin)
def export_reports_excel(request):
    """Export reports data to Excel"""
    # Get the same filter parameters as reports_view
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    room_filter = request.GET.get('room')
    
    # Default to current month if no dates provided
    if not start_date or not end_date:
        today = date.today()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    
    # Parse dates
    try:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        today = date.today()
        start_date_obj = today.replace(day=1)
        end_date_obj = today
    
    # Get the same data as reports_view
    reservations_qs = Reservation.objects.filter(
        date__gte=start_date_obj,
        date__lte=end_date_obj
    )
    
    if room_filter:
        reservations_qs = reservations_qs.filter(room_id=room_filter)
    
    room_stats = reservations_qs.values(
        'room__code'
    ).annotate(
        reservation_count=Count('id')
    ).order_by('room__code')
    
    material_stats = ReservationItem.objects.filter(
        reservation__in=reservations_qs
    ).values(
        'material__name'
    ).annotate(
        total_quantity=Sum('quantity')
    ).order_by('material__name')
    
    # Check if there's data to export
    if not reservations_qs.exists():
        messages.error(request, "No hay datos para exportar en el período seleccionado.")
        return redirect('reports')
    
    # Create Excel workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create summary sheet
    summary_ws = wb.create_sheet("Resumen")
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    summary_ws['A1'] = "Reporte de Biblioteca"
    summary_ws['A1'].font = Font(bold=True, size=16)
    summary_ws.merge_cells('A1:C1')
    
    # Date range
    summary_ws['A3'] = f"Período: {start_date_obj.strftime('%d/%m/%Y')} - {end_date_obj.strftime('%d/%m/%Y')}"
    summary_ws.merge_cells('A3:C3')
    
    # Summary statistics
    summary_ws['A5'] = "Métrica"
    summary_ws['B5'] = "Valor"
    summary_ws['A5'].font = header_font
    summary_ws['A5'].fill = header_fill
    summary_ws['A5'].alignment = header_alignment
    summary_ws['B5'].font = header_font
    summary_ws['B5'].fill = header_fill
    summary_ws['B5'].alignment = header_alignment
    
    summary_ws['A6'] = "Total de reservas"
    summary_ws['B6'] = reservations_qs.count()
    summary_ws['A7'] = "Salones utilizados"
    summary_ws['B7'] = len(room_stats)
    summary_ws['A8'] = "Tipos de materiales"
    summary_ws['B8'] = len(material_stats)
    
    # Adjust column widths
    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 15
    
    # Create room statistics sheet
    room_ws = wb.create_sheet("Reservas por Salón")
    
    # Headers
    room_ws['A1'] = "Código de Salón"
    room_ws['B1'] = "Cantidad de Reservas"
    
    for col in ['A1', 'B1']:
        room_ws[col].font = header_font
        room_ws[col].fill = header_fill
        room_ws[col].alignment = header_alignment
    
    # Data
    row = 2
    for stat in room_stats:
        room_ws[f'A{row}'] = f"Salón {stat['room__code']}"
        room_ws[f'B{row}'] = stat['reservation_count']
        row += 1
    
    # Adjust column widths
    room_ws.column_dimensions['A'].width = 20
    room_ws.column_dimensions['B'].width = 25
    
    # Create material statistics sheet
    material_ws = wb.create_sheet("Materiales Solicitados")
    
    # Headers
    material_ws['A1'] = "Material"
    material_ws['B1'] = "Cantidad Total Solicitada"
    
    for col in ['A1', 'B1']:
        material_ws[col].font = header_font
        material_ws[col].fill = header_fill
        material_ws[col].alignment = header_alignment
    
    # Data
    row = 2
    for stat in material_stats:
        material_ws[f'A{row}'] = stat['material__name']
        material_ws[f'B{row}'] = stat['total_quantity']
        row += 1
    
    # Adjust column widths
    material_ws.column_dimensions['A'].width = 30
    material_ws.column_dimensions['B'].width = 25
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_biblioteca_{start_date}_{end_date}.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response
